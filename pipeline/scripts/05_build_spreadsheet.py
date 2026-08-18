"""Step 5 - Match the four raw sources together and write the Excel workbook.

Matching strategy
-----------------
Disease names almost never match verbatim across sources. Orphanet says
"Cystic fibrosis"; FDA says "Treatment of patients with cystic fibrosis";
ClinicalTrials.gov says "Cystic Fibrosis (CF)". So each disease string is
reduced to a *match key* - lowercased, stripped of boilerplate ("treatment of",
"patients with"), parentheticals and punctuation, with roman numerals folded to
digits and non-discriminating words removed. Keys are compared with rapidfuzz's
token_set_ratio, which tolerates the extra words FDA indication text carries.

Two passes:
  1. exact match on the normalised key (fast, high confidence, scored 100)
  2. fuzzy match at >= MATCH_THRESHOLD for everything left over

Every emitted link carries its match_score and match_method so the result can be
audited or filtered in Excel rather than taken on faith.

Writes output/rare_disease_database.xlsx (+ output/map_full.csv if the map sheet
exceeds Excel's row ceiling).

Run standalone:  python scripts/05_build_spreadsheet.py [--threshold 88]
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))

from common import (  # noqa: E402
    CLINICAL_TRIALS_CSV,
    DISEASES_CSV,
    FDA_APPROVALS_CSV,
    FDA_ORPHAN_CSV,
    MAP_OVERFLOW_CSV,
    STAGE_RANK,
    XLSX_PATH,
    clean_company,
    clean_product,
    content_tokens,
    get_logger,
    most_advanced,
    normalize_name,
)

LOG = get_logger("05_build")

MATCH_THRESHOLD = 88
EXCEL_MAX_ROWS = 1_048_576
MAP_SHEET_CAP = 900_000       # leave headroom under Excel's hard ceiling
CELL_CHAR_LIMIT = 32_000

# Interventions that represent an actual therapeutic product.
PRODUCT_INTERVENTION_TYPES = {
    "DRUG", "BIOLOGICAL", "GENETIC", "COMBINATION_PRODUCT", "DIETARY_SUPPLEMENT",
}

# Intervention names that are controls or procedures, not products.
NON_PRODUCT_RE = re.compile(
    r"^(placebo|saline|normal saline|sham|standard of care|best supportive care|"
    r"no intervention|control|vehicle|water|sugar pill|matching placebo|"
    r"blood draw|skin biopsy|questionnaire|survey|observation|physical therapy|"
    r"exercise|mri|ct scan|electroencephalography|blood sample|placebo comparator)\b",
    re.IGNORECASE,
)

ILLEGAL_XLSX = re.compile(r"[\000-\010\013\014\016-\037]")

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)


# ------------------------------------------------------------------ matching


# Explicit negation carries meaning no string metric can see: "non-Hodgkin
# lymphoma" and "Hodgkin lymphoma" are different diseases, yet token_set_ratio
# scores them 100 because one token set contains the other.
NEGATION_TOKENS = frozenset({"non", "not", "without", "excluding", "negative"})

# Structural guards against token_set_ratio's subset blind spot.
MIN_CONTAINMENT = 0.6     # IDF-weighted share of the disease's tokens that must match
MAX_MISSING_TOKENS = 1    # disease tokens absent from the source text
MIN_SOLO_TOKEN_IDF = 5.5  # a lone shared token must be rare enough to stand alone
TOP_CANDIDATES = 25       # re-rank this many scorer hits structurally

# Inverse document frequency over the disease-name corpus, populated by
# build_disease_index. Without it, a pair sharing only filler words like
# "invasive infections" scores as highly as one sharing "aspergillus".
_IDF: dict[str, float] = {}
_DEFAULT_IDF = 6.0        # unseen token: treat as maximally distinctive


def token_weight(tokens) -> float:
    """Summed IDF of a token set; rare, disease-specific words dominate."""
    return sum(_IDF.get(t, _DEFAULT_IDF) for t in tokens)


def match_key(text: str | None) -> str:
    """Canonical comparison key: sorted, de-noised content tokens."""
    tokens = content_tokens(text)
    return " ".join(sorted(tokens))


def structural_verdict(source: frozenset[str], disease: frozenset[str]):
    """Decide whether two token sets may be linked, and how good the link is.

    Returns (ok, rank_tuple). `source` is the free text (FDA indication or trial
    condition); `disease` is the Orphanet name or synonym.

    token_set_ratio alone accepts any subset relation, which produces links like
    "eosinophilic gastritis" -> "Eosinophilic cystitis". These rules require the
    two strings to genuinely overlap rather than merely share a prefix of tokens.
    """
    if not source or not disease:
        return False, ()

    # Negation must agree on both sides.
    if (source & NEGATION_TOKENS) != (disease & NEGATION_TOKENS):
        return False, ()

    # Identical token sets are an exact match once normalised ("Treatment of
    # gliomas" vs "Glioma"). No heuristic should get a vote on those.
    if source == disease:
        return True, (1.0, len(source), 0, 0)

    shared = source & disease
    missing = disease - source
    if not shared:
        return False, ()

    # A single shared token only counts if it is rare enough to identify a
    # disease on its own. Judge that by IDF, not by character length: "alport"
    # is six letters but names exactly one disorder, while "infections" is ten
    # letters and names hundreds.
    if len(shared) == 1 and token_weight(shared) < MIN_SOLO_TOKEN_IDF:
        return False, ()

    if len(missing) > MAX_MISSING_TOKENS:
        return False, ()

    # Weight by IDF so the overlap has to be carried by distinctive words.
    # "invasive Aspergillus infections" vs "Invasive VRE infections" shares two
    # tokens, but both are common filler, so the weighted containment is low.
    total = token_weight(disease)
    if total <= 0:
        return False, ()
    containment = token_weight(shared) / total
    if containment < MIN_CONTAINMENT:
        return False, ()

    # Prefer the most specific surviving candidate: highest weighted containment,
    # most shared tokens, fewest missing, then closest overall length.
    rank = (round(containment, 3), len(shared), -len(missing), -len(source ^ disease))
    return True, rank


def build_disease_index(diseases: pd.DataFrame):
    """Return (exact_map, choice_keys, choice_orpha, choice_tokens).

    Every primary name and synonym contributes a key. True diseases outrank
    Orphanet's category/group nodes when two entries share a key. The fuzzy
    candidate pool is restricted to real disorders: matching an FDA indication
    onto a bucket like "Rare immune disease" is noise, not a finding.
    """
    exact: dict[str, tuple[str, int]] = {}   # normalised name -> (orpha, priority)
    variants: dict[str, tuple[str, int]] = {}

    for row in diseases.itertuples(index=False):
        orpha = row.orpha_code
        is_disease = row.is_disease == "True"
        names = [row.disease_name]
        if row.synonyms:
            names.extend(s.strip() for s in row.synonyms.split(";"))

        for pos, name in enumerate(names):
            if not name or not name.strip():
                continue
            # Prefer true diseases, then primary names over synonyms.
            priority = (2 if is_disease else 0) + (1 if pos == 0 else 0)

            norm = normalize_name(name)
            prev_exact = exact.get(norm)
            if norm and (prev_exact is None or priority > prev_exact[1]):
                exact[norm] = (orpha, priority)

            if not is_disease:
                continue
            key = match_key(name)
            if not key:
                continue
            # Drop keys with no token long enough to be discriminating; these
            # produce absurd hits (Orphanet's "C syndrome" reduces to "c").
            if max(len(t) for t in key.split()) < 4:
                continue
            prev = variants.get(key)
            if prev is None or priority > prev[1]:
                variants[key] = (orpha, priority)

    choice_keys = list(variants.keys())
    choice_orpha = [variants[k][0] for k in choice_keys]
    choice_tokens = [frozenset(k.split()) for k in choice_keys]

    # Build the IDF table from the disease-name corpus.
    doc_freq: dict[str, int] = defaultdict(int)
    for tokens in choice_tokens:
        for token in tokens:
            doc_freq[token] += 1
    n_docs = max(len(choice_tokens), 1)
    _IDF.clear()
    for token, freq in doc_freq.items():
        _IDF[token] = float(np.log(n_docs / freq))

    return exact, choice_keys, choice_orpha, choice_tokens


def match_texts(
    texts: list[str],
    exact: dict[str, tuple[str, int]],
    choice_keys: list[str],
    choice_orpha: list[str],
    choice_tokens: list[frozenset[str]],
    threshold: int,
) -> dict[str, tuple[str, int, str]]:
    """Map each input string to (orpha_code, score, method). Unmatched are omitted.

    Pass 1 is an exact match on the normalised name. Pass 2 scores the remainder
    with token_set_ratio, then re-ranks the top candidates structurally, because
    the raw score alone accepts too many subset coincidences.
    """
    from rapidfuzz import fuzz, process

    resolved: dict[str, tuple[str, int, str]] = {}
    pending: list[str] = []

    for text in texts:
        norm = normalize_name(text)
        hit = exact.get(norm)
        if hit:
            resolved[text] = (hit[0], 100, "exact")
            continue
        if match_key(text):
            pending.append(text)

    LOG.info("  %s exact matches, %s going to fuzzy pass",
             f"{len(resolved):,}", f"{len(pending):,}")
    if not pending or not choice_keys:
        return resolved

    queries = [match_key(t) for t in pending]
    query_tokens = [frozenset(q.split()) for q in queries]
    chunk = 2000
    n_fuzzy = n_rejected = 0

    for start in range(0, len(queries), chunk):
        block = queries[start:start + chunk]
        scores = process.cdist(
            block, choice_keys,
            scorer=fuzz.token_set_ratio,
            dtype=np.uint8,
            workers=-1,
            score_cutoff=threshold,
        )
        for offset in range(len(block)):
            row = scores[offset]
            hits = np.flatnonzero(row >= threshold)
            if hits.size == 0:
                continue
            if hits.size > TOP_CANDIDATES:
                hits = hits[np.argpartition(row[hits], -TOP_CANDIDATES)[-TOP_CANDIDATES:]]

            q_tokens = query_tokens[start + offset]
            best = None
            for idx in hits:
                ok, rank = structural_verdict(q_tokens, choice_tokens[idx])
                if not ok:
                    continue
                candidate = (rank, int(row[idx]), idx)
                if best is None or candidate > best:
                    best = candidate

            if best is None:
                n_rejected += 1
                continue
            _rank, score, idx = best
            resolved[pending[start + offset]] = (choice_orpha[idx], score, "fuzzy")
            n_fuzzy += 1

        LOG.info("    fuzzy %s/%s ...", f"{min(start + chunk, len(queries)):,}",
                 f"{len(queries):,}")

    LOG.info("  %s fuzzy matches accepted, %s scorer hits rejected by structural rules",
             f"{n_fuzzy:,}", f"{n_rejected:,}")
    return resolved


def verify_trial_conditions(trials: pd.DataFrame, diseases: pd.DataFrame,
                            threshold: int) -> pd.Series:
    """Score each trial row: does the trial's condition text really match the disease?

    Trials were retrieved by querying the disease name, but ClinicalTrials.gov's
    condition search is loose, so a query can return studies about a neighbouring
    disorder. This re-checks each hit against the disease's own name and synonyms.
    """
    from rapidfuzz import fuzz

    names_by_orpha: dict[str, list[tuple[str, frozenset[str]]]] = {}
    for row in diseases.itertuples(index=False):
        keys = [match_key(row.disease_name)]
        if row.synonyms:
            keys.extend(match_key(s) for s in row.synonyms.split(";"))
        names_by_orpha[row.orpha_code] = [
            (k, frozenset(k.split())) for k in keys if k
        ]

    memo: dict[tuple[str, str], int] = {}
    scores = np.zeros(len(trials), dtype=np.uint8)

    orpha_col = trials["orpha_code"].to_numpy()
    cond_col = trials["trial_conditions"].to_numpy()

    for i in range(len(trials)):
        orpha = orpha_col[i]
        conditions = cond_col[i]
        cache_key = (orpha, conditions)
        cached = memo.get(cache_key)
        if cached is not None:
            scores[i] = cached
            continue

        disease_keys = names_by_orpha.get(orpha) or []
        best = 0
        for condition in str(conditions).split(";"):
            ckey = match_key(condition)
            if not ckey:
                continue
            ctokens = frozenset(ckey.split())
            for dkey, dtokens in disease_keys:
                s = fuzz.token_set_ratio(ckey, dkey, score_cutoff=threshold)
                if s <= best:
                    continue
                # Same structural guard as the FDA pass: a high scorer value on
                # its own admits neighbouring-disease trials.
                ok, _rank = structural_verdict(ctokens, dtokens)
                if not ok:
                    continue
                best = int(s)
                if best == 100:
                    break
            if best == 100:
                break
        memo[cache_key] = best
        scores[i] = best

    LOG.info("  verified %s trial rows (%s distinct disease/condition pairs)",
             f"{len(trials):,}", f"{len(memo):,}")
    return pd.Series(scores, index=trials.index)


# ------------------------------------------------------------------ loading


def load_sources() -> tuple[pd.DataFrame, ...]:
    def read(path: Path, label: str) -> pd.DataFrame:
        if not path.exists():
            LOG.warning("%s missing (%s) - continuing without it", label, path.name)
            return pd.DataFrame()
        df = pd.read_csv(path, dtype=str).fillna("")
        LOG.info("Loaded %-22s %s rows", label, f"{len(df):,}")
        return df

    diseases = read(DISEASES_CSV, "diseases")
    orphan = read(FDA_ORPHAN_CSV, "fda orphan")
    approvals = read(FDA_APPROVALS_CSV, "fda approvals")
    trials = read(CLINICAL_TRIALS_CSV, "clinical trials")

    if diseases.empty:
        raise SystemExit("diseases_raw.csv is required - run 01_fetch_diseases.py first.")
    return diseases, orphan, approvals, trials


def build_approval_lookup(approvals: pd.DataFrame) -> dict[str, dict]:
    """Map a normalised product name to its earliest approval record."""
    lookup: dict[str, dict] = {}
    if approvals.empty:
        return lookup

    for row in approvals.itertuples(index=False):
        date = row.approval_date
        info = {
            "approval_date": date,
            "mechanism": row.mechanism_moa or row.pharm_class_epc,
            "company": row.company_name,
            "brand": row.product_brand_name,
            "route": row.route,
            "marketing_status": row.marketing_status,
        }
        candidates = {clean_product(row.product_brand_name),
                      clean_product(row.product_generic_name)}
        for ingredient in str(row.active_ingredients).split(";"):
            candidates.add(clean_product(ingredient))
        for key in candidates:
            if not key or len(key) < 3:
                continue
            prev = lookup.get(key)
            if prev is None or (date and (not prev["approval_date"] or date < prev["approval_date"])):
                lookup[key] = info
    return lookup


# ------------------------------------------------------------------ link building


def links_from_orphan(orphan: pd.DataFrame, matches: dict, approval_lookup: dict) -> list[dict]:
    """One link row per FDA orphan designation whose indication resolved to a disease."""
    links = []
    for row in orphan.itertuples(index=False):
        hit = matches.get(row.designated_indication)
        if not hit:
            continue
        orpha, score, method = hit

        product_display = (row.product_trade_name or row.product_generic_name).strip()
        product_key = clean_product(row.product_trade_name) or clean_product(row.product_generic_name)
        if not product_key:
            continue

        approval = approval_lookup.get(product_key) or approval_lookup.get(
            clean_product(row.product_generic_name))
        # "Approved" here must mean approved *for this indication*, which is what
        # FDA's designation status records. A Drugs@FDA hit only proves the
        # molecule is approved for something, so it cannot promote the stage.
        stage = "Approved" if row.is_approved == "True" else "Orphan Designated"

        links.append({
            "orpha_code": orpha,
            "company_name": row.company_name,
            "company_key": clean_company(row.company_name),
            "product_name": product_display,
            "product_key": product_key,
            "product_generic": row.product_generic_name,
            "development_stage": stage,
            "source": "FDA Orphan Designation",
            "evidence": f"Designated {row.designation_date}" if row.designation_date else "",
            "source_indication": row.designated_indication,
            "match_score": score,
            "match_method": method,
            "trial_status": "",
            "nct_id": "",
            "approval_date": (approval or {}).get("approval_date", "") if approval else "",
            "mechanism": (approval or {}).get("mechanism", "") if approval else "",
            "approved_any_indication": "Yes" if approval else "No",
        })
    return links


def links_from_trials(trials: pd.DataFrame, approval_lookup: dict) -> list[dict]:
    """One link row per verified trial intervention."""
    links = []
    for row in trials.itertuples(index=False):
        product_display = row.intervention_name.strip()
        product_key = clean_product(product_display)
        if not product_key or len(product_key) < 3:
            continue

        approval = approval_lookup.get(product_key)
        # The trial phase is the evidence for *this* disease. A drug approved
        # for an unrelated indication is still only Phase N here - showing
        # "Approved" would imply it is a therapy for this disease.
        stage = row.phase

        company = row.lead_sponsor
        links.append({
            "orpha_code": row.orpha_code,
            "company_name": company,
            "company_key": clean_company(company),
            "product_name": product_display,
            "product_key": product_key,
            "product_generic": "",
            "development_stage": stage,
            "source": "ClinicalTrials.gov",
            "evidence": row.nct_id,
            "source_indication": row.trial_conditions,
            "match_score": int(row.condition_score),
            "match_method": "trial-verified",
            "trial_status": row.overall_status,
            "nct_id": row.nct_id,
            "approval_date": (approval or {}).get("approval_date", ""),
            "mechanism": (approval or {}).get("mechanism", ""),
            "approved_any_indication": "Yes" if approval else "No",
        })
    return links


def collapse_links(links: list[dict]) -> pd.DataFrame:
    """Dedupe to one row per (disease, company, product), keeping the best evidence."""
    if not links:
        return pd.DataFrame()

    merged: dict[tuple[str, str, str], dict] = {}
    for link in links:
        key = (link["orpha_code"], link["company_key"], link["product_key"])
        existing = merged.get(key)
        if existing is None:
            link = dict(link)
            link["sources"] = {link["source"]}
            link["nct_ids"] = {link["nct_id"]} if link["nct_id"] else set()
            merged[key] = link
            continue

        existing["sources"].add(link["source"])
        if link["nct_id"]:
            existing["nct_ids"].add(link["nct_id"])
        existing["development_stage"] = most_advanced(
            [existing["development_stage"], link["development_stage"]])
        if link["match_score"] > existing["match_score"]:
            existing["match_score"] = link["match_score"]
            existing["match_method"] = link["match_method"]
        if link.get("approved_any_indication") == "Yes":
            existing["approved_any_indication"] = "Yes"
        for field in ("approval_date", "mechanism", "product_generic", "trial_status"):
            if not existing.get(field) and link.get(field):
                existing[field] = link[field]

    rows = []
    for link in merged.values():
        link["source"] = " + ".join(sorted(link["sources"]))
        ncts = sorted(link["nct_ids"])
        link["n_trials"] = len(ncts)
        link["evidence"] = (
            "; ".join(ncts[:8]) + (f" (+{len(ncts) - 8} more)" if len(ncts) > 8 else "")
        ) if ncts else link["evidence"]
        link.pop("sources", None)
        link.pop("nct_ids", None)
        rows.append(link)

    return pd.DataFrame(rows)


# ------------------------------------------------------------------ sheets


def build_sheets(diseases, orphan, approvals, trials, link_df):
    disease_meta = diseases.set_index("orpha_code")

    # Trial counts come from the verified trial rows, not from summing link
    # counts: one study testing three drugs is three links but one trial.
    def distinct_trials(column: str) -> pd.Series:
        if trials.empty or column not in trials.columns:
            return pd.Series(dtype=int)
        return trials.groupby(column)["nct_id"].nunique()

    trials_by_disease = distinct_trials("orpha_code")
    trials_by_company = distinct_trials("company_key")
    trials_by_product = distinct_trials("product_key")

    # ---- aggregate per disease
    if not link_df.empty:
        by_disease = link_df.groupby("orpha_code").agg(
            n_companies=("company_key", "nunique"),
            n_products=("product_key", "nunique"),
            n_links=("product_key", "size"),
        )
        stage_by_disease = link_df.groupby("orpha_code")["development_stage"].apply(most_advanced)
    else:
        by_disease = pd.DataFrame(columns=["n_companies", "n_products", "n_links"])
        stage_by_disease = pd.Series(dtype=str)

    orphan_counts = pd.Series(dtype=int)
    if not orphan.empty and "matched_orpha" in orphan.columns:
        orphan_counts = orphan[orphan["matched_orpha"] != ""].groupby("matched_orpha").size()

    diseases_sheet = pd.DataFrame({
        "Orpha Code": disease_meta.index,
        "Disease Name": disease_meta["disease_name"],
        "Category": disease_meta["category"],
        "All Categories": disease_meta["all_categories"],
        "Record Type": np.where(disease_meta["is_disease"] == "True", "Disease",
                                disease_meta["disorder_group"]),
        "Disorder Type": disease_meta["disorder_type"],
        "Prevalence": disease_meta["prevalence_class"],
        "Prevalence Type": disease_meta["prevalence_type"],
        "Prevalence Region": disease_meta["prevalence_geographic"],
        "Synonyms": disease_meta["synonyms"],
        "ICD-10": disease_meta["icd10_code"],
        "OMIM": disease_meta["omim_code"],
        "Description": disease_meta["definition"],
        "Orphanet URL": disease_meta["orphanet_url"],
    }).reset_index(drop=True)

    diseases_sheet["Companies"] = diseases_sheet["Orpha Code"].map(
        by_disease["n_companies"] if not by_disease.empty else {}).fillna(0).astype(int)
    diseases_sheet["Products"] = diseases_sheet["Orpha Code"].map(
        by_disease["n_products"] if not by_disease.empty else {}).fillna(0).astype(int)
    diseases_sheet["Trials"] = diseases_sheet["Orpha Code"].map(
        trials_by_disease).fillna(0).astype(int)
    diseases_sheet["FDA Designations"] = diseases_sheet["Orpha Code"].map(
        orphan_counts).fillna(0).astype(int)
    diseases_sheet["Most Advanced Stage"] = diseases_sheet["Orpha Code"].map(
        stage_by_disease).fillna("None found")

    diseases_sheet = diseases_sheet[[
        "Orpha Code", "Disease Name", "Category", "Record Type", "Prevalence",
        "Companies", "Products", "Trials", "FDA Designations", "Most Advanced Stage",
        "Description", "Synonyms", "ICD-10", "OMIM", "Disorder Type",
        "Prevalence Type", "Prevalence Region", "All Categories", "Orphanet URL",
    ]].sort_values(["Companies", "Disease Name"], ascending=[False, True])

    # ---- companies
    if not link_df.empty:
        name_by_key = (link_df.groupby("company_key")["company_name"]
                       .agg(lambda s: s.value_counts().idxmax()))
        companies_sheet = link_df.groupby("company_key").agg(
            Diseases=("orpha_code", "nunique"),
            Products=("product_key", "nunique"),
            Links=("product_key", "size"),
        ).reset_index()
        companies_sheet["Trials"] = companies_sheet["company_key"].map(
            trials_by_company).fillna(0).astype(int)
        companies_sheet["Company Name"] = companies_sheet["company_key"].map(name_by_key)
        companies_sheet["Most Advanced Stage"] = companies_sheet["company_key"].map(
            link_df.groupby("company_key")["development_stage"].apply(most_advanced))
        companies_sheet["Has Approved Product"] = np.where(
            companies_sheet["Most Advanced Stage"] == "Approved", "Yes", "No")
        companies_sheet["Sources"] = companies_sheet["company_key"].map(
            link_df.groupby("company_key")["source"].apply(
                lambda s: " + ".join(sorted({p for v in s for p in v.split(" + ")}))))
        companies_sheet = companies_sheet[[
            "Company Name", "Diseases", "Products", "Links", "Trials",
            "Most Advanced Stage", "Has Approved Product", "Sources",
        ]].sort_values(["Products", "Diseases"], ascending=False)
    else:
        companies_sheet = pd.DataFrame(columns=["Company Name", "Diseases", "Products"])

    # ---- products
    if not link_df.empty:
        disp_by_key = (link_df.groupby("product_key")["product_name"]
                       .agg(lambda s: s.value_counts().idxmax()))
        products_sheet = link_df.groupby("product_key").agg(
            Diseases=("orpha_code", "nunique"),
            Companies=("company_key", "nunique"),
        ).reset_index()
        products_sheet["Trials"] = products_sheet["product_key"].map(
            trials_by_product).fillna(0).astype(int)
        products_sheet["Product Name"] = products_sheet["product_key"].map(disp_by_key)
        products_sheet["Development Stage"] = products_sheet["product_key"].map(
            link_df.groupby("product_key")["development_stage"].apply(most_advanced))
        products_sheet["Mechanism"] = products_sheet["product_key"].map(
            link_df.groupby("product_key")["mechanism"].apply(
                lambda s: next((v for v in s if v), "")))
        products_sheet["Approval Date"] = products_sheet["product_key"].map(
            link_df.groupby("product_key")["approval_date"].apply(
                lambda s: next((v for v in s if v), "")))
        products_sheet["FDA Approved (any indication)"] = products_sheet["product_key"].map(
            link_df.groupby("product_key")["approved_any_indication"].apply(
                lambda s: "Yes" if (s == "Yes").any() else "No"))
        products_sheet["Companies (names)"] = products_sheet["product_key"].map(
            link_df.groupby("product_key")["company_name"].apply(
                lambda s: "; ".join(sorted(set(s))[:5])))
        products_sheet["Sources"] = products_sheet["product_key"].map(
            link_df.groupby("product_key")["source"].apply(
                lambda s: " + ".join(sorted({p for v in s for p in v.split(" + ")}))))
        products_sheet = products_sheet[[
            "Product Name", "Development Stage", "Mechanism", "Diseases", "Companies",
            "Trials", "FDA Approved (any indication)", "Approval Date",
            "Companies (names)", "Sources",
        ]].sort_values(["Diseases", "Product Name"], ascending=[False, True])
    else:
        products_sheet = pd.DataFrame(columns=["Product Name", "Development Stage"])

    # ---- master map
    if not link_df.empty:
        map_sheet = link_df.copy()
        map_sheet["Disease Name"] = map_sheet["orpha_code"].map(disease_meta["disease_name"])
        map_sheet["Category"] = map_sheet["orpha_code"].map(disease_meta["category"])
        map_sheet["Prevalence"] = map_sheet["orpha_code"].map(disease_meta["prevalence_class"])
        map_sheet = map_sheet.rename(columns={
            "orpha_code": "Orpha Code",
            "company_name": "Company Name",
            "product_name": "Product Name",
            "development_stage": "Development Stage",
            "mechanism": "Mechanism",
            "source": "Source",
            "evidence": "Evidence",
            "match_score": "Match Score",
            "match_method": "Match Method",
            "trial_status": "Trial Status",
            "approval_date": "Approval Date",
            "n_trials": "Trials",
            "source_indication": "Source Text Matched",
            "approved_any_indication": "Drug Approved Elsewhere",
        })
        map_sheet = map_sheet[[
            "Disease Name", "Orpha Code", "Category", "Company Name", "Product Name",
            "Development Stage", "Mechanism", "Source", "Trials", "Trial Status",
            "Drug Approved Elsewhere", "Approval Date", "Evidence", "Match Score",
            "Match Method", "Prevalence", "Source Text Matched",
        ]]
        stage_rank = map_sheet["Development Stage"].map(STAGE_RANK).fillna(0)
        map_sheet = map_sheet.assign(_rank=stage_rank).sort_values(
            ["_rank", "Match Score", "Disease Name"], ascending=[False, False, True]
        ).drop(columns="_rank")
    else:
        map_sheet = pd.DataFrame(columns=["Disease Name", "Company Name", "Product Name"])

    return diseases_sheet, companies_sheet, products_sheet, map_sheet


# ------------------------------------------------------------------ excel writing


def _clean_cell(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, (int, float, np.integer, np.floating)):
        return value.item() if hasattr(value, "item") else value
    text = ILLEGAL_XLSX.sub("", str(value))
    return text[:CELL_CHAR_LIMIT]


def write_workbook(sheets: dict[str, tuple[pd.DataFrame, dict[str, int]]]) -> None:
    wb = Workbook(write_only=True)

    for name, (df, widths) in sheets.items():
        ws = wb.create_sheet(title=name)
        ws.freeze_panes = "A2"

        for i, column in enumerate(df.columns, start=1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(column, 18)

        header = []
        for column in df.columns:
            from openpyxl.cell import WriteOnlyCell
            cell = WriteOnlyCell(ws, value=str(column))
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(vertical="center", horizontal="left")
            header.append(cell)
        ws.append(header)

        for record in df.itertuples(index=False, name=None):
            ws.append([_clean_cell(v) for v in record])

        LOG.info("  sheet %-28s %s rows x %d cols", name, f"{len(df):,}", len(df.columns))

    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX_PATH)


WIDTHS = {
    "Disease Name": 46, "Category": 34, "Description": 80, "Synonyms": 46,
    "All Categories": 52, "Orphanet URL": 46, "Company Name": 40,
    "Product Name": 38, "Mechanism": 46, "Source": 30, "Evidence": 34,
    "Source Text Matched": 60, "Companies (names)": 46, "Sources": 30,
    "Development Stage": 20, "Most Advanced Stage": 22, "Prevalence": 18,
    "Record Type": 16, "Disorder Type": 24, "Prevalence Type": 20,
    "Prevalence Region": 20, "Trial Status": 20, "Match Method": 15,
    "Has Approved Product": 20, "ICD-10": 16, "OMIM": 14, "Orpha Code": 12,
    "Drug Approved Elsewhere": 22, "FDA Approved (any indication)": 26,
}


# ------------------------------------------------------------------ main


def build(threshold: int = MATCH_THRESHOLD) -> None:
    diseases, orphan, approvals, trials = load_sources()

    LOG.info("Indexing disease names and synonyms ...")
    exact, choice_keys, choice_orpha, choice_tokens = build_disease_index(diseases)
    LOG.info("  %s fuzzy candidate keys (true disorders only) from %s Orphanet entries",
             f"{len(choice_keys):,}", f"{len(diseases):,}")

    approval_lookup = build_approval_lookup(approvals)
    LOG.info("Built approval lookup with %s product keys", f"{len(approval_lookup):,}")

    links: list[dict] = []

    if not orphan.empty:
        LOG.info("Matching FDA orphan indications to Orphanet diseases ...")
        indications = sorted(set(orphan["designated_indication"]) - {""})
        matches = match_texts(indications, exact, choice_keys, choice_orpha,
                              choice_tokens, threshold)
        LOG.info("  %s/%s distinct indications matched (%.1f%%)",
                 f"{len(matches):,}", f"{len(indications):,}",
                 100 * len(matches) / max(len(indications), 1))
        orphan["matched_orpha"] = orphan["designated_indication"].map(
            lambda t: matches.get(t, ("", 0, ""))[0])
        links.extend(links_from_orphan(orphan, matches, approval_lookup))
        LOG.info("  %s links from FDA designations", f"{len(links):,}")

    if not trials.empty:
        LOG.info("Verifying ClinicalTrials.gov condition matches ...")
        trials = trials[trials["intervention_name"] != ""].copy()
        keep_type = trials["intervention_type"].isin(PRODUCT_INTERVENTION_TYPES)
        not_control = ~trials["intervention_name"].str.match(NON_PRODUCT_RE)
        trials = trials[keep_type & not_control].copy()
        LOG.info("  %s rows remain after keeping therapeutic interventions",
                 f"{len(trials):,}")

        trials["condition_score"] = verify_trial_conditions(trials, diseases, threshold)
        before = len(trials)
        trials = trials[trials["condition_score"] >= threshold].copy()
        LOG.info("  %s/%s rows passed condition verification at >= %d (%.1f%%)",
                 f"{len(trials):,}", f"{before:,}", threshold,
                 100 * len(trials) / max(before, 1))

        n_before = len(links)
        links.extend(links_from_trials(trials, approval_lookup))
        LOG.info("  %s links from clinical trials", f"{len(links) - n_before:,}")

        # Keys for the distinct-trial counts in build_sheets.
        trials["company_key"] = trials["lead_sponsor"].map(clean_company)
        trials["product_key"] = trials["intervention_name"].map(clean_product)

    LOG.info("Collapsing %s raw links to unique disease/company/product ...", f"{len(links):,}")
    link_df = collapse_links(links)
    LOG.info("  %s unique links", f"{len(link_df):,}")

    LOG.info("Building sheets ...")
    diseases_sheet, companies_sheet, products_sheet, map_sheet = build_sheets(
        diseases, orphan, approvals, trials, link_df)

    # Excel row ceiling safety valve: full table always goes to CSV as well.
    if len(map_sheet) > MAP_SHEET_CAP:
        LOG.warning("Map has %s rows, above the %s cap - writing full table to %s",
                    f"{len(map_sheet):,}", f"{MAP_SHEET_CAP:,}", MAP_OVERFLOW_CSV.name)
        map_sheet.to_csv(MAP_OVERFLOW_CSV, index=False, encoding="utf-8")
        map_sheet = map_sheet.head(MAP_SHEET_CAP)
    else:
        MAP_OVERFLOW_CSV.unlink(missing_ok=True)

    LOG.info("Writing %s ...", XLSX_PATH.name)
    write_workbook({
        "Diseases": (diseases_sheet, WIDTHS),
        "Companies": (companies_sheet, WIDTHS),
        "Products": (products_sheet, WIDTHS),
        "Disease-Company-Product Map": (map_sheet, WIDTHS),
    })

    LOG.info("=" * 62)
    LOG.info("Wrote %s", XLSX_PATH)
    LOG.info("  Diseases   %s rows (%s with >=1 company)",
             f"{len(diseases_sheet):,}", f"{int((diseases_sheet['Companies'] > 0).sum()):,}")
    LOG.info("  Companies  %s rows", f"{len(companies_sheet):,}")
    LOG.info("  Products   %s rows", f"{len(products_sheet):,}")
    LOG.info("  Map        %s rows", f"{len(map_sheet):,}")
    if not products_sheet.empty:
        LOG.info("  stage mix: %s",
                 products_sheet["Development Stage"].value_counts().to_dict())
    LOG.info("=" * 62)


def main() -> None:
    parser = argparse.ArgumentParser(description="Match sources and build the Excel workbook.")
    parser.add_argument("--threshold", type=int, default=MATCH_THRESHOLD,
                        help=f"Fuzzy match cutoff 0-100 (default {MATCH_THRESHOLD}).")
    args = parser.parse_args()
    build(threshold=args.threshold)


if __name__ == "__main__":
    main()
